from django.shortcuts import render
from django.contrib import messages

import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from geopy.geocoders import Nominatim

# API_KEY = AIzaSyBqujzL0RDaOJEPjX-1tTuBDXbk58u3H9g

# from geopy.geocoders import GoogleV3
# geolocator = GoogleV3()
# location = geolocator.geocode("175 5th Avenue NYC")
# print(location.address)
# print((location.latitude, location.longitude))
geolocator = Nominatim(user_agent='geoadd')


def index(request):
    data = {}
    if "GET" == request.method:
        return render(request, 'index.html', {})
    # if not GET then proceed
    else:
        excel_file = request.FILES["excel_file"]
        # if not excel_file.name.endswith('.xlsx'):
        #     messages.error(request, 'File is not Excel type')
        #     return HttpResponseRedirect(reverse("Getaddress:excel_upload"))
        #
        # # check file size
        # if excel_file.multiple_chunks():
        #     messages.error(request, "Uploaded file is too big (%.2f MB)." % (excel_file.size / (1000 * 1000),))
        #     return HttpResponseRedirect(reverse("Getaddress:excel_upload"))

        # file_data = excel_file.read().decode("utf-8")

        wb = openpyxl.load_workbook(excel_file)
        wba = Workbook(write_only=True)
        ws = wba.create_sheet()

        # getting all sheets
        sheets = wb.sheetnames
        # print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        # print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        cols = active_sheet.max_column

        # getting max_row in excel sheet
        rows = active_sheet.max_row

        # reading a cell
        # print(worksheet["A1"].value)

        excel_data = list()
        lat_data = list()
        lng_data = list()
        # iterating over the rows and

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                location = geolocator.geocode(cell.value)
                print(cell.value)
                row_data.append(str(location.latitude))
                row_data.append(str(location.longitude))
                row_data.append(str(cell.value))
                # print(cell.value)
            excel_data.append(row_data)

        if excel_data:
            for line in excel_data:
                ws.append(line)

        wba.save('address.xlsx')

        return render(request, 'index.html', {"excel_data": excel_data})
